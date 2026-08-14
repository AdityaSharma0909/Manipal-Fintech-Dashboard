from __future__ import annotations

import json
from io import BytesIO
from typing import Any, Iterable, Optional

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from onboarding_v2.models import BankLeadTrace, PincodeMaster


BANK_CRM_REPORT_HEADERS = [
    "Timestamp",
    "FirstName",
    "LastName",
    "MobileNo",
    "PinCode",
    "LoanAmount",
    "CrmId",
    "CrmType",
    "Bank Name",
    "SO ID",
    "CreatedOn",
    "Api Response Message",
    "State",
    "District",
    "BranchCode",
]


def _safe_dict(value: Any) -> dict:
    return value if isinstance(value, dict) else {}


def _first_value(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return ""


def _split_name(full_name: Optional[str]) -> tuple[str, str]:
    parts = [part for part in str(full_name or "").strip().split() if part]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _local_datetime(value) -> str:
    if not value:
        return ""
    return timezone.localtime(value).strftime("%Y-%m-%d %H:%M:%S")


def _json_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, default=str, ensure_ascii=False)
    return str(value)


def _mask_mobile(value: Any) -> str:
    mobile = str(value or "").strip()
    if len(mobile) <= 4:
        return mobile
    return f"{'*' * (len(mobile) - 4)}{mobile[-4:]}"


def _response_message(trace: BankLeadTrace) -> str:
    response = _safe_dict(trace.response_payload)
    decrypted = response.get("decrypted_body")
    if decrypted:
        return _json_text(decrypted)

    json_body = response.get("json_body")
    if json_body:
        return _json_text(json_body)

    raw = response.get("raw_body")
    if isinstance(raw, str):
        try:
            raw_obj = json.loads(raw)
            if isinstance(raw_obj, (dict, list)):
                return _json_text(raw_obj)
        except json.JSONDecodeError:
            pass
        return raw

    if response:
        return _json_text(response)

    return str(trace.error_message or "")


def _location_from_pincode(pincode: Any) -> tuple[str, str]:
    if not pincode:
        return "", ""

    record = PincodeMaster.objects.filter(pincode=str(pincode).strip()).first()
    if not record:
        return "", ""
    return record.statename or "", record.district or ""


def _trace_to_row(trace: BankLeadTrace) -> list[Any]:
    lead = trace.lead
    metadata = _safe_dict(trace.metadata)
    onboarding_payload = _safe_dict(metadata.get("onboarding_request_payload"))
    lead_metadata = _safe_dict(getattr(lead, "metadata", None))

    full_name = _first_value(
        getattr(lead, "customer_name", None),
        onboarding_payload.get("customer_name"),
        onboarding_payload.get("FullName"),
    )
    first_name, last_name = _split_name(full_name)

    created_by = trace.created_by or getattr(lead, "created_by", None)
    so_id = _first_value(
        getattr(created_by, "employee_id", None),
        getattr(created_by, "pk", None),
    )

    pincode = _first_value(getattr(lead, "pincode", None), onboarding_payload.get("pincode"), onboarding_payload.get("PinCode"))
    pincode_state, pincode_district = _location_from_pincode(pincode)
    state = _first_value(lead_metadata.get("state"), onboarding_payload.get("state"), onboarding_payload.get("SBOState"), pincode_state)
    district = _first_value(lead_metadata.get("district"), onboarding_payload.get("district"), onboarding_payload.get("SBODistrict"), pincode_district)
    crm_type = _first_value(trace.crm_type, getattr(lead, "crm_type", None), lead_metadata.get("crm_type"), onboarding_payload.get("crm_type"))

    return [
        _local_datetime(trace.created_at),
        first_name,
        last_name,
        _mask_mobile(_first_value(trace.contact_number, getattr(lead, "contact_number", None), onboarding_payload.get("contact_number"))),
        pincode,
        _first_value(getattr(lead, "amount", None), onboarding_payload.get("amount"), onboarding_payload.get("LoanAmount")),
        _first_value(trace.bank_lead_id, getattr(lead, "BankLeadID", None), "NULL"),
        crm_type,
        _first_value(trace.bank_name, getattr(lead, "bank", None), onboarding_payload.get("bank")),
        so_id,
        _local_datetime(getattr(lead, "created_at", None) or trace.created_at),
        _response_message(trace),
        state,
        district,
        _first_value(getattr(lead, "bank_branch", None), onboarding_payload.get("branch_code"), onboarding_payload.get("BranchCode"), onboarding_payload.get("Branch")),
    ]


def build_bank_crm_report_workbook(traces: Iterable[BankLeadTrace]) -> tuple[BytesIO, int]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    sheet.append(BANK_CRM_REPORT_HEADERS)
    count = 0
    for trace in traces:
        sheet.append(_trace_to_row(trace))
        count += 1

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = [20, 18, 18, 15, 12, 15, 18, 16, 18, 14, 20, 45, 18, 18, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer, count
