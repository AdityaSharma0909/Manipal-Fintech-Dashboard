import datetime
import traceback
import pandas as pd
from io import BytesIO as IO # for modern python
from django.http import HttpResponse as dhttp
from rest_framework.views import APIView
from utils.constants import ROLES
from utils.responseHandler import HttpResponse
from application.services.cam_report import ExportCamService
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment


class ExportCamView(APIView):

    def get(self, request):
        try:
            user = request.user
            if user.role != ROLES.CPC.value and user.role != ROLES.CREDIT_MANAGER.value and user.role != ROLES.BUSINESS_HEAD.value:
                return HttpResponse.BadRequest("Not Allowed")

            base_cam_data, reference_data_list, credit_data, bank_detail_list , luc_data ,score_me_data , pd_tele_data , policy_deviation_data , disbursement_note_data = ExportCamService().get_cam_data(request)

            # Creating DataFrames for each section
            personal_info_df = pd.DataFrame([base_cam_data], columns=['Name', 'Email', 'Phone', 'Address','Loan ID','Nature of Business','Business Vintage in Years', 'Nominee'])
            reference_pd_df = pd.DataFrame(reference_data_list, columns=['Reference Name', 'Business', 'Residential', 'No of Years', 'No of Family Members', 'Nature of Business', 'Sub Nature of Business', 'Earning Members', 'Phone', 'Relationship with Applicant'])
            credit_status_df = pd.DataFrame([credit_data], columns=['House Ownership', 'House Number of Years', 'Shop Ownership', 'Shop Number of Years', 'Nature of Business', 'Monthly Income', 'Monthly Expenditure', 'No of Loans Running', 'No of Loans Closed in Last 1 Year', 'Any Loan Applied in Last 30 Days', 'Account Held for No of Years', 'Fixed Assets Held by Him and Family'])
            bank_details_df = pd.DataFrame([bank_detail_list], columns=['Bank Name', 'Account Number', 'IFSC'])
            luc_df = pd.DataFrame([luc_data], columns=['Purpose of Loan', 'How much is the Requirement', 'What is the Expected Amount Increase', 'How to Verify the Usage'])
            score_me_df = pd.DataFrame([score_me_data], columns=['CB Score', 'Obligation', 'Cash Flow Monthly','AMB of 6 Months','Existing Loan Amount','EMI of Existing loan','Leverage to Income', 'Other Source of Income'])
            pd_tele_df = pd.DataFrame([pd_tele_data], columns=['PD Report in Brief', 'PD Done By', 'Tele PD Done by', 'Location Captured', 'Pictures Captured of House/shop', 'Observation', 'Observation Comment', 'Residential Stability', 'Residential Stability Comment', 'Business Stability', 'Business Stability Comment',  'No of Similar Business in the Area', 'External Income', 'Suppliers/Customers Feedback'])
            policy_deviation_df = pd.DataFrame([policy_deviation_data], columns=['Product Name', 'Eligible Amount', 'Deviation Amount','Approved By','Recommended By','Reason for Deviation','Reason for approval'])
            
            disbursement_note_df = pd.DataFrame([disbursement_note_data], columns=["Name of customer","Name of Sales Person","Name of Supervisor","Product name","Deviation if any","Approval for deviation","Loan Amount","Tenure","PF","Insurance value","Cross sell","Net amount to be disbursed","Bank account No","IFSC Code","Penny Drop","First Party only","Bank Name","Bank Branch", "Radian Branch","Authorised by","Recommended By","Processed by","Approved By"])
            excel_file = IO()
            xlwriter = pd.ExcelWriter(excel_file, engine='openpyxl')

            workbook = xlwriter.book
            worksheet = workbook.create_sheet('Cam Report')

            # Define borders
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
            
            # Define font and alignment
            bold_font = Font(bold=True)
            left_alignment = Alignment(horizontal='left', vertical='center')

            # Function to apply border to a range of cells
            def apply_border_to_range(worksheet, range_str, border):
                for row in worksheet[range_str]:
                    for cell in row:
                        cell.border = border

            # Adding headers and merging cells for Personal Information
            worksheet.merge_cells('B2:I2')
            worksheet['B2'] = 'PERSONAL INFORMATION'
            worksheet['B2'].font = bold_font
            worksheet['B2'].alignment = left_alignment
            apply_border_to_range(worksheet, 'B2:I10', thick_border)  # Apply border from B2 to I10

            row = 3
            for col, value in personal_info_df.items():
                worksheet[f'B{row}'] = col
                worksheet[f'B{row}'].border = thick_border
                worksheet[f'B{row}'].alignment = left_alignment
                worksheet[f'C{row}'] = value.iloc[0]
                worksheet[f'C{row}'].border = thick_border
                worksheet[f'C{row}'].alignment = left_alignment
                row += 1

            start_row = 13
            worksheet.merge_cells(f'B{start_row}:I{start_row}')
            worksheet[f'B{start_row}'] = 'CREDIT REPORT'
            worksheet[f'B{start_row}'].font = bold_font
            worksheet[f'B{start_row}'].alignment = left_alignment
            apply_border_to_range(worksheet, f'B{start_row}:I21', thick_border)  # Apply border from B13 to I21

            report_row = start_row + 1
            for col, value in score_me_df.items():
                worksheet[f'B{report_row}'] = col
                worksheet[f'B{report_row}'].border = thick_border
                worksheet[f'B{report_row}'].alignment = left_alignment
                worksheet[f'C{report_row}'] = value.iloc[0]
                worksheet[f'C{report_row}'].border = thick_border
                worksheet[f'C{report_row}'].alignment = left_alignment
                report_row += 1

            # Writing headers for Reference Details
            start_row = 24
            worksheet.merge_cells(f'B{start_row}:I{start_row}')  # Merge B to I for headers
            worksheet[f'B{start_row}'] = 'REFERENCE DETAILS'
            worksheet[f'B{start_row}'].font = bold_font
            worksheet[f'B{start_row}'].alignment = left_alignment
            apply_border_to_range(worksheet, f'B{start_row}:I34', thick_border)  # Apply border from B24 to I34

            # Writing headers in the first column starting from B, C, etc.
            row = start_row + 1
            for header in reference_pd_df.columns:
                worksheet[f'B{row}'] = header
                worksheet[f'B{row}'].border = thick_border
                worksheet[f'B{row}'].alignment = left_alignment
                row += 1

            # Now, write the values horizontally for each header
            row = start_row + 1
            for col_index in range(len(reference_pd_df)):
                col = 3 + 2 * col_index  # Start from C (3rd column), then E (5th column), and so on
                row = start_row + 1
                for header in reference_pd_df.columns:
                    value = reference_pd_df.at[col_index, header]  # Fetching value from DataFrame
                    worksheet[f'{chr(64 + col)}{row}'] = value
                    worksheet[f'{chr(64 + col)}{row}'].border = thick_border
                    worksheet[f'{chr(64 + col)}{row}'].alignment = left_alignment
                    row += 1

            # Adding headers and merging cells for Credit Status
            start_row = 37
            worksheet.merge_cells(f'B{start_row}:I{start_row}')
            worksheet[f'B{start_row}'] = 'CREDIT STATUS'
            worksheet[f'B{start_row}'].font = bold_font
            worksheet[f'B{start_row}'].alignment = left_alignment
            apply_border_to_range(worksheet, f'B{start_row}:I50', thick_border)  # Apply border from B37 to I50

            credit_row = start_row + 1
            for col, value in credit_status_df.items():
                worksheet[f'B{credit_row}'] = col
                worksheet[f'B{credit_row}'].border = thick_border
                worksheet[f'B{credit_row}'].alignment = left_alignment
                worksheet[f'C{credit_row}'] = value.iloc[0]
                worksheet[f'C{credit_row}'].border = thick_border
                worksheet[f'C{credit_row}'].alignment = left_alignment
                credit_row += 1

            # Adding headers and values for Bank Details
            bank_details_start_row = credit_row 
            worksheet[f'B{bank_details_start_row}'] = 'Bank Details'
            worksheet[f'B{bank_details_start_row}'].border = thick_border
            worksheet[f'B{bank_details_start_row}'].alignment = left_alignment
            apply_border_to_range(worksheet, f'B{bank_details_start_row}:E{bank_details_start_row}', thick_border)  # Apply border from B{bank_details_start_row} to E{bank_details_start_row}

            worksheet[f'C{bank_details_start_row}'] = bank_details_df.iloc[0]['Bank Name']
            worksheet[f'C{bank_details_start_row}'].border = thick_border
            worksheet[f'C{bank_details_start_row}'].alignment = left_alignment
            worksheet[f'D{bank_details_start_row}'] = bank_details_df.iloc[0]['Account Number']
            worksheet[f'D{bank_details_start_row}'].border = thick_border
            worksheet[f'D{bank_details_start_row}'].alignment = left_alignment
            worksheet[f'E{bank_details_start_row}'] = bank_details_df.iloc[0]['IFSC']
            worksheet[f'E{bank_details_start_row}'].border = thick_border
            worksheet[f'E{bank_details_start_row}'].alignment = left_alignment

            start_row = 53
            worksheet.merge_cells(f'B{start_row}:I{start_row}')
            worksheet[f'B{start_row}'] = 'LUC'
            worksheet[f'B{start_row}'].font = bold_font
            worksheet[f'B{start_row}'].alignment = left_alignment
            apply_border_to_range(worksheet, f'B{start_row}:I57', thick_border)  # Apply border from B53 to I57

            luc_row = start_row + 1
            for col, value in luc_df.items():
                worksheet[f'B{luc_row}'] = col
                worksheet[f'B{luc_row}'].border = thick_border
                worksheet[f'B{luc_row}'].alignment = left_alignment
                worksheet[f'C{luc_row}'] = value.iloc[0]
                worksheet[f'C{luc_row}'].border = thick_border
                worksheet[f'C{luc_row}'].alignment = left_alignment
                luc_row += 1

            start_row = 60
            worksheet.merge_cells(f'B{start_row}:I{start_row}')
            worksheet[f'B{start_row}'] = 'PD & TELE PD'
            worksheet[f'B{start_row}'].font = bold_font
            worksheet[f'B{start_row}'].alignment = left_alignment
            apply_border_to_range(worksheet, f'B{start_row}:I74', thick_border)  # Apply border from B60 to I72

            pd_row = start_row + 1
            for col, value in pd_tele_df.items():
                worksheet[f'B{pd_row}'] = col
                worksheet[f'B{pd_row}'].border = thick_border
                worksheet[f'B{pd_row}'].alignment = left_alignment
                worksheet[f'C{pd_row}'] = value.iloc[0]
                worksheet[f'C{pd_row}'].border = thick_border
                worksheet[f'C{pd_row}'].alignment = left_alignment
                pd_row += 1

            start_row = 77
            worksheet.merge_cells(f'B{start_row}:I{start_row}')
            worksheet[f'B{start_row}'] = 'POLICY & DEVIATION'
            worksheet[f'B{start_row}'].font = bold_font
            worksheet[f'B{start_row}'].alignment = left_alignment
            apply_border_to_range(worksheet, f'B{start_row}:I84', thick_border)  # Apply border from B75 to I82

            policy_row = start_row + 1
            for col, value in policy_deviation_df.items():
                worksheet[f'B{policy_row}'] = col
                worksheet[f'B{policy_row}'].border = thick_border
                worksheet[f'B{policy_row}'].alignment = left_alignment
                worksheet[f'C{policy_row}'] = value.iloc[0]
                worksheet[f'C{policy_row}'].border = thick_border
                worksheet[f'C{policy_row}'].alignment = left_alignment
                policy_row += 1
            
            # Create a new sheet for the disbursement note
            disbursement_sheet = workbook.create_sheet(title="Disbursement Note")

            # Write Disbursement Note data to the new sheet
            start_row = 2
            disbursement_sheet.merge_cells(f'B{start_row}:D{start_row}')
            disbursement_sheet[f'B{start_row}'] = 'DISBURSEMENT NOTE'
            disbursement_sheet[f'B{start_row}'].font = bold_font
            disbursement_sheet[f'B{start_row}'].alignment = left_alignment
            apply_border_to_range(disbursement_sheet, f'B{start_row}:D{start_row + 23}', thick_border)  # Apply border from B1 to I26

            row = start_row + 1
            for col, value in disbursement_note_df.items():
                disbursement_sheet[f'B{row}'] = col
                disbursement_sheet[f'B{row}'].border = thick_border
                disbursement_sheet[f'B{row}'].alignment = left_alignment
                disbursement_sheet[f'C{row}'] = value.iloc[0]
                disbursement_sheet[f'C{row}'].border = thick_border
                disbursement_sheet[f'C{row}'].alignment = left_alignment
                row += 1


            xlwriter.close()
            excel_file.seek(0)

            response = dhttp(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            application_number = base_cam_data[4]
            current_date = datetime.datetime.now().strftime('%Y-%m-%d')
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Disposition'] = f'attachment; filename=Cam_Report_{application_number}_{current_date}.xlsx'
            return response

        except Exception as e:
            traceback.print_exc()
            return HttpResponse.InternalServerError(str(e))