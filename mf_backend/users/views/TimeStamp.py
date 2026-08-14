from rest_framework.views import APIView
from ..serializers import TimeStampSerializer
from utils.responseHandler import HttpResponse
import traceback
from users.models import User
from utils.constants import TIMESTAMP, ROLES
import datetime
import io
import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from users.models import TimeStamp
from utility.e2e_utility import MinioUtility
from django.conf import settings
from django.core.mail import EmailMessage
from utils.envSetup import environment
from django.utils import timezone
from django.http import HttpResponse as DjangoHttpResponse
from ..tasks import export_today_timestamps_task
from users.selfie_urls import get_selfie_access_url
import pytz


class TimeStampView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            user = request.user
            data = request.data
            status_to_save = None
            try:
                if user.role not in [ROLES.SALES_OFFICER.value, ROLES.REGIONAL_HEAD.value]:
                    return HttpResponse.Forbidden(
                        "Sales Officers and Regional Heads are only allowed to check in"
                    )
                
                lastTimestamp = TimeStamp.objects.filter(
                    user=user, created_at__gte=datetime.date.today()
                ).latest("created_at")

                if lastTimestamp.status == TIMESTAMP.CHECKED_IN.value:
                    status_to_save = TIMESTAMP.CHECKED_OUT.value
                elif lastTimestamp.status == TIMESTAMP.CHECKED_OUT.value:
                    status_to_save = TIMESTAMP.CHECKED_IN.value
                else:
                    return HttpResponse.InternalServerError("Something went wrong!!!")

            except TimeStamp.DoesNotExist as te:
                status_to_save = TIMESTAMP.CHECKED_IN.value

            selfie = request.FILES.get("selfie")
            selfie_url = None
            if selfie:
                try:
                    selfie_url = MinioUtility().put_objects(file=selfie, path=settings.SELFIE)
                except Exception as e:
                    return HttpResponse.InternalServerError(f"E2E upload failed: {str(e)}")

            data_to_serialize = {
                "user": str(user.user_id),
                "status": status_to_save,
                "latitude": data.get("latitude"),
                "longitude": data.get("longitude"),
                "remarks": data.get("remarks"),
            }

            serializer = TimeStampSerializer(data=data_to_serialize)

            if serializer.is_valid():
                timestamp_obj = serializer.save()
                if selfie_url:
                    timestamp_obj.selfie = selfie_url
                    timestamp_obj.save()
                return HttpResponse.Success({"timeStamp": TimeStampSerializer(timestamp_obj).data})
            else:
                print("timestampSer.errors::: ", serializer.errors)
                return HttpResponse.BadRequest(serializer.errors)
        except Exception as e:
            traceback.print_exc()
            return HttpResponse.InternalServerError(str(e))


class TimeStampExportView(APIView):
    def get(self, request, *args, **kwargs):
        try:
            result = export_today_timestamps_task()
            
            if "exported and sent" in result:
                return HttpResponse.Success(result)
            elif "No records found" in result:
                return HttpResponse.Success(result)
            else:
                return HttpResponse.InternalServerError(result)

        except Exception as e:
            traceback.print_exc()
            return HttpResponse.InternalServerError(str(e))


class TimeStampDownloadExcelView(APIView):
    """
    Download Styled TimeStamp records as an Excel (.xlsx) file with native hyperlinks.
    """

    def get(self, request, *args, **kwargs):
        try:
            IST = pytz.timezone('Asia/Kolkata')

            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')
            date_str = request.query_params.get('date')

            if start_date_str and end_date_str:
                try:
                    start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
                    end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
                except ValueError:
                    return HttpResponse.BadRequest("Invalid date format. Use YYYY-MM-DD.")
            elif date_str:
                try:
                    start_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
                    end_date = start_date
                except ValueError:
                    return HttpResponse.BadRequest("Invalid date format. Use YYYY-MM-DD.")
            else:
                start_date = datetime.date.today()
                end_date = start_date

            timestamps = TimeStamp.objects.filter(
                created_at__date__gte=start_date,
                created_at__date__lte=end_date,
            ).select_related('user').order_by('created_at')

            status_filter = request.query_params.get('status')
            if status_filter:
                timestamps = timestamps.filter(status=status_filter)

            employee_id_filter = request.query_params.get('employee_id')
            if employee_id_filter:
                timestamps = timestamps.filter(user__employee_id=employee_id_filter)

            if not timestamps.exists():
                return HttpResponse.BadRequest("No records found for the selected date range.")

            # ── Data Rows ──────────────────────────────────────────
            rows = []
            for idx, ts in enumerate(timestamps, start=1):
                user = ts.user
                full_name = f"{user.first_name or ''} {user.last_name or ''}".strip() if user else "N/A"
                ist_time = ts.created_at.astimezone(IST) if ts.created_at else None

                rows.append({
                    "S.No": idx,
                    "Emp ID": user.employee_id if user else "N/A",
                    "User Type": user.role if user else "N/A",
                    "Team": user.team if user and user.team else "N/A",
                    "User Name": full_name,
                    "User Mobile No": str(user.phone) if user else "N/A",
                    "Status": ts.status,
                    "Latitude": ts.latitude,
                    "Longitude": ts.longitude,
                    "User Work Log Time": ist_time.strftime("%d-%m-%Y %H:%M:%S") if ist_time else "N/A",
                    "Server Work Log Timestamp": ts.created_at.strftime("%d-%m-%Y %H:%M:%S") if ts.created_at else "N/A",
                    "Image URL": get_selfie_access_url(ts.selfie) if ts.selfie else "No Image",
                    "Location URL": f"https://maps.google.com/?q={ts.latitude},{ts.longitude}" if ts.latitude and ts.longitude else "N/A",
                    "Remarks": ts.remarks or "",
                })

            df = pd.DataFrame(rows)
            excel_buffer = io.BytesIO()
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Attendance Report', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Attendance Report']

                # Styles
                green_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
                white_font = Font(color='FFFFFF', bold=True)
                link_font = Font(color='0563C1', underline='single')
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                center_alignment = Alignment(horizontal='center', vertical='center')

                for cell in worksheet[1]:
                    cell.fill = green_fill
                    cell.font = white_font
                    cell.border = thin_border
                    cell.alignment = center_alignment

                # Clickable native hyperlinks
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='center')
                        
                        # Columns L (Image) and M (Location) are links in this view
                        if cell.column_letter in ['L', 'M'] and cell.value and cell.value.startswith('http'):
                            url = cell.value
                            cell.value = "Click to View" if cell.column_letter == 'L' else "View on Map"
                            cell.hyperlink = url
                            cell.font = link_font

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    worksheet.column_dimensions[column].width = max_length + 2

            excel_buffer.seek(0)
            
            if start_date == end_date:
                filename = f"Attendance_Report_{start_date.strftime('%d-%m-%Y')}.xlsx"
            else:
                filename = f"Attendance_Report_{start_date.strftime('%d-%m-%Y')}_{end_date.strftime('%d-%m-%Y')}.xlsx"

            response = DjangoHttpResponse(
                excel_buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            traceback.print_exc()
            return HttpResponse.InternalServerError(str(e))
