from celery import shared_task
import datetime
import io
import traceback
import pytz
import pandas as pd
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from users.models import TimeStamp
from users.selfie_urls import get_selfie_access_url
from utils.envSetup import environment
from utils.constants import TIMESTAMP
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

@shared_task(name='export_today_timestamps_task')
def export_today_timestamps_task(recipient_email=None):
    """
    Exports today's TimeStamp records to a professional styled Excel file (.xlsx) with clickable links.
    """
    try:
        IST = pytz.timezone('Asia/Kolkata')
        today = datetime.date.today()
        
        # ── Fetch Data ──────────────────────────────────────────────
        timestamps = TimeStamp.objects.filter(created_at__date=today).select_related('user').order_by("-created_at")

        if not timestamps.exists():
            return "No records found for today to export."

        # ── Calculate Summary Stats ─────────────────────────────────
        total_count = timestamps.count()
        check_in_count = timestamps.filter(status=TIMESTAMP.CHECKED_IN.value).count()
        check_out_count = timestamps.filter(status=TIMESTAMP.CHECKED_OUT.value).count()

        # ── Prepare Data Rows ───────────────────────────────────────
        rows = []
        for ts in timestamps:
            user = ts.user
            full_name = f"{user.first_name or ''} {user.last_name or ''}".strip() if user else "N/A"
            ist_time = ts.created_at.astimezone(IST) if ts.created_at else None
            
            rows.append({
                "Emp ID": user.employee_id if user else "N/A",
                "User Type": user.role if user else "N/A",
                "Team": user.team if user and user.team else "N/A",
                "User Name": full_name,
                "User Mobile No": str(user.phone) if user else "N/A",
                "Status": ts.status,
                "Latitude": ts.latitude,
                "Longitude": ts.longitude,
                "User Work Log Time": ist_time.strftime("%d-%m-%Y %H:%M:%S") if ist_time else "N/A",
                "Server Work Log Timestamp": ts.created_at.strftime("%d-%m-%Y %H:%M:%S") if ts.created_at else "N/A",
                "Image URL": get_selfie_access_url(ts.selfie) if ts.selfie else "No Image",
                "Location URL": f"https://maps.google.com/?q={ts.latitude},{ts.longitude}" if ts.latitude and ts.longitude else "N/A",
                "Remarks": ts.remarks or ""
            })

        # ── Generate & Style Excel ─────────────────────────────────
        df = pd.DataFrame(rows)
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance Report', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Attendance Report']

            # Styles
            green_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
            white_font = Font(color='FFFFFF', bold=True)
            link_font = Font(color='0563C1', underline='single')  # Standard Excel link blue
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')

            # Headers
            for cell in worksheet[1]:
                cell.fill = green_fill
                cell.font = white_font
                cell.border = thin_border
                cell.alignment = center_alignment

            # Data Cells and Clickable Links
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                    
                    # Columns K (Image) and L (Location) are links
                    if cell.column_letter in ['K', 'L'] and cell.value and cell.value.startswith('http'):
                        url = cell.value
                        cell.value = "Click to View" if cell.column_letter == 'K' else "View on Map"
                        cell.hyperlink = url
                        cell.font = link_font

            # Auto-adjust column widths
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column].width = max_length + 2

        excel_buffer.seek(0)
        excel_content = excel_buffer.getvalue()

        # ── Email Context & Sending ────────────────────────────────
        context = {
            'report_date': today.strftime("%d %B, %Y"),
            'total_count': total_count,
            'check_in_count': check_in_count,
            'check_out_count': check_out_count,
        }

        html_content = render_to_string('users/email/timestamp_report.html', context)
        text_content = f"Today's SO Logging Report ({today})\n\nTotal: {total_count}\nCheck-ins: {check_in_count}\nCheck-outs: {check_out_count}"

        raw_recipients = recipient_email or environment.TIMESTAMP_EXPORT_EMAIL
        if not raw_recipients:
            return "Recipient email not set in environment variables."

        # Convert comma-separated string to a clean list of emails
        if isinstance(raw_recipients, str):
            recipient_list = [e.strip() for e in raw_recipients.split(",") if e.strip()]
        else:
            recipient_list = raw_recipients if isinstance(raw_recipients, list) else [raw_recipients]

        subject = f"Daily SO Logging Report - {today.strftime('%d-%m-%Y')}"
        email = EmailMultiAlternatives(
            subject=subject,
            body=text_content,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=recipient_list,
        )
        email.attach_alternative(html_content, "text/html")
        email.attach(
            f"so_logging_report_{today}.xlsx",
            excel_content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        email.send()

        return f"Today's report has been exported and sent to {', '.join(recipient_list)}"

    except Exception as e:
        traceback.print_exc()
        return str(e)
